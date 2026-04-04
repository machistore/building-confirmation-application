"""建築確認申請書 生成メインスクリプト

使い方:
    python src/generator.py [YAMLファイルパス]

    YAMLファイルパスを省略した場合は input/sample_project.yaml を使用する。
"""

import sys
from pathlib import Path

# src/ から実行された場合も import できるよう親ディレクトリをパスに追加
sys.path.insert(0, str(Path(__file__).parent))

import yaml
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy
from calculator import calc_total_floor_area, calc_kenpei_ratio, calc_yoseki_ratio
from pdf_converter import convert_to_pdf
from validator import validate


def load_yaml(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def format_result(data: dict, total_floor_area: float, kenpei: float, yoseki: float) -> str:
    lines = []
    meta = data.get("meta", {})
    site = data.get("敷地", {})

    lines.append("=" * 50)
    lines.append("建築確認申請書 計算結果")
    lines.append("=" * 50)
    lines.append(f"案件番号  : {meta.get('案件番号', '---')}")
    lines.append(f"担当者    : {meta.get('担当者', '---')}")
    lines.append("")
    lines.append("【敷地情報】")
    lines.append(f"  敷地面積      : {site.get('敷地面積', '---')} ㎡")
    lines.append(f"  指定建蔽率    : {site.get('指定建蔽率', '---')} %")
    lines.append(f"  指定容積率    : {site.get('指定容積率', '---')} %")
    lines.append("")
    lines.append("【建築面積・延べ床面積】")
    lines.append(f"  建築面積      : {data.get('建築面積', '---')} ㎡")
    lines.append(f"  延べ床面積    : {total_floor_area} ㎡")
    lines.append("")
    lines.append("【各階床面積】")
    for floor in data.get("各階", []):
        lines.append(f"  {floor['階']:<6}: {floor['床面積']} ㎡")
    lines.append("")
    lines.append("【計算結果】")
    lines.append(f"  建蔽率        : {kenpei} %  （指定: {site.get('指定建蔽率', '---')} %）")

    kenpei_ok = kenpei <= site.get("指定建蔽率", float("inf"))
    lines.append(f"  建蔽率チェック: {'OK' if kenpei_ok else 'NG - 指定建蔽率を超過しています'}")

    lines.append(f"  容積率        : {yoseki} %  （指定: {site.get('指定容積率', '---')} %）")
    yoseki_ok = yoseki <= site.get("指定容積率", float("inf"))
    lines.append(f"  容積率チェック: {'OK' if yoseki_ok else 'NG - 指定容積率を超過しています'}")

    lines.append("=" * 50)
    return "\n".join(lines)


def _resolve_key(data: dict, dotted_key: str):
    """ドット区切りキーでネストした辞書の値を取得する。

    例: "建築主.氏名" → data["建築主"]["氏名"]
    """
    keys = dotted_key.split(".")
    value = data
    for k in keys:
        if not isinstance(value, dict):
            return None
        value = value.get(k)
    return value


def write_excel(
    data: dict,
    calc_values: dict,
    cell_map_path: Path,
    output_path: Path,
) -> None:
    """cell_map.yaml に従ってデータを確認用 xlsx に書き込む（1シート・フラット構成）。

    cell_map の各エントリを読み、source/key に従って値を取得し
    openpyxl の ws.cell(row, column) で書き込む（row/col ともに 0始まり→1始まりに変換）。

    Args:
        data: YAMLから読み込んだ入力データ
        calc_values: 計算済み値の辞書（延べ床面積・建蔽率・容積率）
        cell_map_path: cell_map.yaml のパス
        output_path: 出力先 .xlsx のパス
    """
    cell_map = load_yaml(cell_map_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "確認申請書"

    floors_cfg = None

    for entry_name, entry in cell_map.items():
        if not isinstance(entry, dict):
            continue
        if entry_name == "floors":
            floors_cfg = entry
            continue
        if "row" not in entry or "col" not in entry:
            continue

        source = entry.get("source")
        key = entry.get("key", "")
        if source == "data":
            value = _resolve_key(data, key)
        elif source == "calc":
            value = calc_values.get(key)
        else:
            continue

        if value is None:
            continue

        # cell_map は 0始まり → openpyxl は 1始まり
        ws.cell(row=entry["row"] + 1, column=entry["col"] + 1, value=value)

    # 階別床面積の繰り返し書き込み
    if floors_cfg:
        start_row = floors_cfg["start_row"]  # 0始まり
        kai_col = floors_cfg["階_col"]        # 0始まり
        area_col = floors_cfg["床面積_col"]   # 0始まり
        for i, floor in enumerate(data.get("各階", [])):
            row = start_row + i
            ws.cell(row=row + 1, column=kai_col + 1, value=floor.get("階"))
            ws.cell(row=row + 1, column=area_col + 1, value=floor.get("床面積"))

    wb.save(output_path)


def write_to_template(
    data: dict,
    calc_values: dict,
    template_path: Path,
    cell_map_path: Path,
    output_path: Path,
) -> None:
    """テンプレート XLS に cell_map.yaml の座標へ値を書き込んで保存する。

    Args:
        data: YAMLから読み込んだ入力データ
        calc_values: 計算済み値の辞書（延べ床面積・建蔽率・容積率）
        template_path: 元テンプレート .xls のパス
        cell_map_path: cell_map.yaml のパス
        output_path: 出力先 .xls のパス
    """
    cell_map = load_yaml(cell_map_path)

    rb = xlrd.open_workbook(str(template_path), formatting_info=True)
    wb = xl_copy(rb)

    floors_cfg = cell_map.pop("floors", None)
    floor_blocks_cfg = cell_map.pop("floor_blocks_第五面", None)
    dokuritsu_cfg = cell_map.pop("独立部分_第六面", None)

    for entry_name, entry in cell_map.items():
        if not isinstance(entry, dict):
            continue
        if "sheet_idx" not in entry or "row" not in entry or "col" not in entry:
            continue

        # 書き込む値を決定
        source = entry.get("source")
        key = entry.get("key", "")
        if source == "data":
            value = _resolve_key(data, key)
        elif source == "calc":
            value = calc_values.get(key)
        else:
            continue

        if value is None:
            continue

        ws = wb.get_sheet(entry["sheet_idx"])
        ws.write(entry["row"], entry["col"], value)

    # 第四面: 階別床面積の繰り返し書き込み
    if floors_cfg and isinstance(floors_cfg, dict):
        ws = wb.get_sheet(floors_cfg["sheet_idx"])
        start_row = floors_cfg["start_row"]
        kai_col = floors_cfg["階_col"]
        area_col = floors_cfg["床面積_col"]
        for i, floor in enumerate(data.get("各階", [])):
            row = start_row + i
            ws.write(row, kai_col, floor.get("階", ""))
            ws.write(row, area_col, floor.get("床面積", ""))

    # 第五面: 建築物の階別概要（ブロック繰り返し）
    if floor_blocks_cfg and isinstance(floor_blocks_cfg, dict):
        _write_floor_blocks(wb, floor_blocks_cfg, data.get("階別概要", []))

    # 第六面: 建築物独立部分別概要
    if dokuritsu_cfg and isinstance(dokuritsu_cfg, dict):
        _write_independent_parts(wb, dokuritsu_cfg, data.get("建築物独立部分", []))

    wb.save(str(output_path))


def _write_floor_blocks(wb, cfg: dict, data_list: list) -> None:
    """第五面: ブロック構造への階別データ書き込み。

    各ブロックは start_row + block_idx * block_stride 行から始まる。
    fields 内の各エントリで row_offset / col / key / type を参照して書き込む。
    type が "checkbox_true"  → 値が真のとき "■"、偽のとき "□"
    type が "checkbox_false" → 値が偽のとき "■"、真のとき "□"
    """
    ws = wb.get_sheet(cfg["sheet_idx"])
    start_row = cfg["start_row"]
    block_stride = cfg["block_stride"]
    fields = cfg.get("fields", {})

    for block_idx, floor_data in enumerate(data_list):
        block_start = start_row + block_idx * block_stride
        for field_name, field_cfg in fields.items():
            row = block_start + field_cfg["row_offset"]
            col = field_cfg["col"]
            key = field_cfg.get("key", field_name)
            value = floor_data.get(key)
            field_type = field_cfg.get("type", "value")

            if field_type == "checkbox_true":
                ws.write(row, col, "■" if value else "□")
            elif field_type == "checkbox_false":
                ws.write(row, col, "■" if not value else "□")
            elif value is not None:
                ws.write(row, col, value)


def _write_independent_parts(wb, cfg: dict, data_list: list) -> None:
    """第六面: 建築物独立部分の書き込み（現在は先頭1件のみ対応）。

    構造フィールドに transform: strip_zou が指定された場合、
    末尾の「造」を除去して書き込む（テンプレートの L9=造 と重複回避）。
    """
    if not data_list:
        return

    ws = wb.get_sheet(cfg["sheet_idx"])
    part_data = data_list[0]
    fields = cfg.get("fields", {})

    for field_name, field_cfg in fields.items():
        row = field_cfg["row"]
        col = field_cfg["col"]
        value = part_data.get(field_name)
        if value is None:
            continue

        if field_cfg.get("transform") == "strip_zou" and isinstance(value, str):
            value = value.removesuffix("造")

        ws.write(row, col, value)


def main():
    # 入力ファイルパスの決定
    base_dir = Path(__file__).parent.parent
    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
    else:
        input_path = base_dir / "input" / "sample_project.yaml"

    print(f"入力ファイル: {input_path}")

    # YAML読み込み
    try:
        data = load_yaml(input_path)
    except FileNotFoundError:
        print(f"エラー: ファイルが見つかりません: {input_path}")
        sys.exit(1)
    except yaml.YAMLError as e:
        print(f"エラー: YAMLの解析に失敗しました: {e}")
        sys.exit(1)

    # バリデーション
    errors = validate(data)
    if errors:
        print("バリデーションエラーが発生しました:")
        for err in errors:
            print(f"  - {err}")
        sys.exit(1)

    # 計算
    floors = data["各階"]
    building_area = data["建築面積"]
    site_area = data["敷地"]["敷地面積"]

    total_floor_area = calc_total_floor_area(floors)
    kenpei = calc_kenpei_ratio(building_area, site_area)
    yoseki = calc_yoseki_ratio(total_floor_area, site_area)

    # 結果をテキストで出力
    result_text = format_result(data, total_floor_area, kenpei, yoseki)
    print(result_text)

    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)

    # テキスト出力
    txt_path = output_dir / "result.txt"
    txt_path.write_text(result_text, encoding="utf-8")
    print(f"\n結果を保存しました: {txt_path}")

    # result.xlsx 出力（旧フォーマット・確認用）
    cell_map_path = output_dir / "cell_map.yaml"
    xlsx_path = output_dir / "result.xlsx"
    write_excel(
        data=data,
        calc_values={
            "延べ床面積": total_floor_area,
            "建蔽率": kenpei,
            "容積率": yoseki,
        },
        cell_map_path=cell_map_path,
        output_path=xlsx_path,
    )
    print(f"Excelを保存しました: {xlsx_path}")

    # テンプレート XLS への書き込み
    template_path = base_dir / "templates" / "BPR003_260323.xls"
    xls_path = output_dir / "申請書_出力.xls"
    write_to_template(
        data=data,
        calc_values={
            "延べ床面積": total_floor_area,
            "建蔽率": kenpei,
            "容積率": yoseki,
        },
        template_path=template_path,
        cell_map_path=cell_map_path,
        output_path=xls_path,
    )
    print(f"申請書テンプレートを保存しました: {xls_path}")

    # XLS → PDF 変換
    try:
        import time
        t0 = time.time()
        pdf_path = convert_to_pdf(str(xls_path))
        elapsed = time.time() - t0
        pdf_size = Path(pdf_path).stat().st_size
        print(f"PDF出力完了: {pdf_path}")
        print(f"  ファイルサイズ: {pdf_size:,} bytes ({pdf_size / 1024:.1f} KB)")
        print(f"  変換時間: {elapsed:.1f} 秒")
    except Exception as e:
        print(f"PDF変換失敗（XLSはそのまま保存済み）: {e}")


if __name__ == "__main__":
    main()
